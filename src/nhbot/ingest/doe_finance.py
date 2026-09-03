"""NH DOE-25 finance layer — expenditure-by-function + revenue-by-source.

Source: the "District Profile" sheet of each district's DOE-25 Annual Financial
Report (one .xlsx per district per year, downloaded from iPlatform). That sheet
is DOE's own district-level rollup: current expenditure by function with dollars
and percent, and total revenue by source with dollars and percent. This is the
"where does the money go" dataset and the honest basis for the admin-share view.

Inputs (data/raw/doe25/<fyYYYY>/*.xlsx):
  filenames encode the DOE district id + year, e.g. Amherst_2025_d017017.xlsx
  -> district_id 17, year 2025.

Outputs (data/processed/):
  nh_district_expenditure.csv  district_id, year, function_code, function_name, amount, pct
  nh_district_revenue.csv      district_id, year, source_code,  source_name,  amount, pct

Everything joins to school_district on district_id (already loaded).

The parser locates the two tables by scanning for their "Function" headers and
reading until the matching total line, so it tolerates row shifts between years.
"""
import csv, re, glob, os
import openpyxl
from nhbot.config import RAW_DIR, PROCESSED_DIR

DOE25_DIR = RAW_DIR / "doe25"
PROFILE_SHEET = "District Profile"

def norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "")).strip()

def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = norm(v).replace("$", "").replace(",", "").replace("%", "")
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def is_code(a):
    """A DOE function/source code like 1100, 2300&2800, 1600-1800,2750, 5310+5390."""
    a = norm(a)
    return bool(a) and bool(re.match(r"^[0-9]", a)) and not a.lower().startswith("total")

def parse_ids(path):
    """(district_id, year) from a filename like Amherst_2025_d017017.xlsx."""
    b = os.path.basename(path)
    m_year = re.search(r"_(\d{4})_", b) or re.search(r"(\d{4})", b)
    m_did = re.search(r"d(\d{3})\d*", b)
    year = int(m_year.group(1)) if m_year else None
    did = int(m_did.group(1)) if m_did else None
    return did, year

CPP_LEVELS = {"elementary": "cpp_elementary", "middle/junior": "cpp_middle",
              "middle": "cpp_middle", "high": "cpp_high", "district total": "cpp_total"}

def parse_profile(ws):
    """Return (expenditures, revenues, cpp) — the two function/source lists of
    {code,name,amount,pct}, plus a cpp dict of per-level cost-per-pupil. Rows are
    read by section; totals/blank rows skipped. Works on read-only worksheets."""
    rows = [list(r[:4]) for r in ws.iter_rows(values_only=True)]  # cols A-D

    exp, rev = [], []
    cpp = {}
    section = None  # None -> (pre-tables, where cpp lives) | 'exp' | 'rev'
    for a, b, c, d in rows:
        na, nb = norm(a), norm(b)
        low = (na + " " + nb).lower()
        if na.lower() == "function":
            section = "rev" if "revenue" in low else "exp"
            continue
        if section is None:
            # per-pupil block sits above the first Function header:
            # label in col B ("Elementary"/"District Total"), value in col C
            key = CPP_LEVELS.get(nb.lower())
            if key and num(c) is not None:
                cpp.setdefault(key, num(c))
            continue
        label = na or nb
        if label.lower().startswith("total"):
            continue
        amount, pct = num(c), num(d)   # $ in col C, % in col D
        if amount is None:
            continue
        # coded rows: code in A + name in B; code-less rows: name in B, A empty
        if is_code(a):
            code, name = na, nb
        else:
            code, name = "", (nb or na)
        (rev if section == "rev" else exp).append(
            {"code": code, "name": name, "amount": amount, "pct": pct})
    return exp, rev, cpp

def synth_codes(rows, prefix):
    """Ensure every row has a non-empty, unique code (PK safety)."""
    seen, out = set(), []
    for i, r in enumerate(rows):
        code = r["code"] or (prefix + re.sub(r"[^a-z0-9]+", "_", r["name"].lower())[:24] or f"{prefix}{i}")
        base = code; n = 1
        while code in seen:
            n += 1; code = f"{base}_{n}"
        seen.add(code)
        out.append({**r, "code": code})
    return out

def main():
    files = sorted(glob.glob(str(DOE25_DIR / "**" / "*.xlsx"), recursive=True))
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    exp_rows, rev_rows, cpp_rows = [], [], []
    n_ok = n_skip = 0
    skipped = []
    for path in files:
        did, year = parse_ids(path)
        if did is None or year is None:
            skipped.append(os.path.basename(path) + " (no id/year)"); n_skip += 1; continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as ex:
            skipped.append(os.path.basename(path) + f" (unreadable: {type(ex).__name__})")
            n_skip += 1; continue
        try:
            if PROFILE_SHEET not in wb.sheetnames:
                skipped.append(os.path.basename(path) + " (no District Profile)"); n_skip += 1; continue
            exp, rev, cpp = parse_profile(wb[PROFILE_SHEET])
        except Exception as ex:
            skipped.append(os.path.basename(path) + f" (parse error: {type(ex).__name__})")
            n_skip += 1; continue
        finally:
            wb.close()
        exp = synth_codes(exp, "E_"); rev = synth_codes(rev, "R_")
        for e in exp:
            exp_rows.append({"district_id": did, "year": year, "function_code": e["code"],
                             "function_name": e["name"], "amount": e["amount"], "pct": e["pct"]})
        for v in rev:
            rev_rows.append({"district_id": did, "year": year, "source_code": v["code"],
                             "source_name": v["name"], "amount": v["amount"], "pct": v["pct"]})
        if cpp:
            cpp_rows.append({"district_id": did, "year": year,
                             "cpp_elementary": cpp.get("cpp_elementary"),
                             "cpp_middle": cpp.get("cpp_middle"),
                             "cpp_high": cpp.get("cpp_high"),
                             "cpp_total": cpp.get("cpp_total")})
        n_ok += 1

    def write(name, rows, cols):
        with open(PROCESSED_DIR / name, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(rows)

    write("nh_district_expenditure.csv", exp_rows,
          ["district_id", "year", "function_code", "function_name", "amount", "pct"])
    write("nh_district_revenue.csv", rev_rows,
          ["district_id", "year", "source_code", "source_name", "amount", "pct"])
    write("nh_district_cpp.csv", cpp_rows,
          ["district_id", "year", "cpp_elementary", "cpp_middle", "cpp_high", "cpp_total"])

    years = sorted({r["year"] for r in exp_rows})
    print(f"DOE-25 finance: parsed {n_ok} files, skipped {n_skip}; years {years[:3]}..{years[-3:]}")
    if skipped:
        print("  skipped:", skipped[:10], "..." if len(skipped) > 10 else "")
    print(f"  expenditure rows: {len(exp_rows)}  revenue rows: {len(rev_rows)}  cpp rows: {len(cpp_rows)}")

if __name__ == "__main__":
    main()
