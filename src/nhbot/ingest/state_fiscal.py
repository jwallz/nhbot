"""State of New Hampshire fiscal data — budget (LBA HB 1 Excel) + revenue
(DAS Monthly Revenue Focus, June year-end PDFs).

Drop the source files in data/raw/state/ and run `nhbot state-fiscal`:
  * the enacted operating-budget Excel  (LBA "HB 1 - ... .xlsx")
  * one or more year-end revenue PDFs   ("FY####_Monthly_Revenue_June.pdf")

Writes three processed CSVs consumed by nhbot load:
  nh_state_budget.csv    fiscal_year, category, department, amount   (TYPE=E lines)
  nh_state_funding.csv   fiscal_year, source, amount                 (TYPE=F lines)
  nh_state_revenue.csv   fiscal_year, source, actual_musd, plan_musd (General+Ed funds)

Budget years come from the two "LEG" columns in the workbook (the biennium).
Revenue years come from each PDF (its page-3 year-to-date "comparison to plan"
table, whose Total-Actual/Total-Plan columns are found by the arithmetic
Actual - Plan = variance, which is layout-robust across the report's quirks).
"""
import csv, re, glob, collections
from pathlib import Path
from nhbot.config import RAW_DIR, PROCESSED_DIR

SRC = RAW_DIR / "state"

# --- department / category name cleanup ------------------------------------
_EXPAND = {
    'Svcs': 'Services', 'Svc': 'Service', 'Prtn': 'Protection', 'Protect': 'Protection',
    'Developmt': 'Development', 'Resrcs': 'Resources', 'Affrs': 'Affairs',
    'Admin': 'Administration', 'Std': 'Standards', 'Stds': 'Standards',
    'Agricult': 'Agriculture', 'Cncl': 'Council', 'Rel': 'Relations',
    'Cert': 'Certification', 'Xfers': 'Transfers', 'Prof': 'Professional',
    'Econ': 'Economic', 'Bus': 'Business', 'Vet': 'Veterans', 'Info': 'Information',
}
_ACR = {'Nh': 'NH', 'It': 'IT', 'Hhs': 'HHS', 'Dhhs': 'DHHS'}
_LOWER = {'and', 'of', 'the', 'for', 'to', 'a', 'an', 'in'}

def _clean_name(s):
    if not s:
        return s
    out = []
    for i, w in enumerate(re.split(r'\s+', s.strip().lower())):
        if w == '&':
            out.append('&'); continue
        core = w.strip('&')
        exp = _EXPAND.get(core.capitalize())
        if exp:
            w = exp
        elif core in _LOWER and i != 0:
            w = core
        else:
            w = core.capitalize()
            w = _ACR.get(w, w)
        out.append(w)
    return ' '.join(out)


def parse_budget(xlsx_path):
    """(budget_rows, funding_rows, federal_rows) from the HB 1 workbook.
    budget rows aggregate TYPE=E expenditure lines by (category, department, FY);
    funding rows aggregate TYPE=F source-of-funds lines by (class name, FY);
    federal rows break the single 'Federal Funds' class down by receiving agency
    (category, department, FY) so the page can show where federal money actually goes."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    # locate the two LEG (fiscal-year) amount columns from the header
    yr_cols = [(i, int(re.search(r'(\d{4})', str(h)).group(1)))
               for i, h in enumerate(hdr) if h and str(h).upper().startswith('LEG')]
    budg = collections.defaultdict(lambda: collections.defaultdict(float))
    fund = collections.defaultdict(lambda: collections.defaultdict(float))
    fed = collections.defaultdict(lambda: collections.defaultdict(float))
    for r in it:
        cat, dept, typ, clsname = r[1], r[3], r[10], r[12]
        if dept is None:
            continue
        for ci, yr in yr_cols:
            amt = r[ci] or 0
            if typ == 'E':
                budg[(_clean_name(cat), _clean_name(dept))][yr] += amt
            elif typ == 'F':
                fund[clsname][yr] += amt
                if str(clsname) == 'Federal Funds':
                    fed[(_clean_name(cat), _clean_name(dept))][yr] += amt
    years = [y for _, y in yr_cols]
    brows = [(yr, cat, dept, round(v[yr])) for (cat, dept), v in budg.items() for yr in years]
    frows = [(yr, src, round(v[yr])) for src, v in fund.items() for yr in years]
    fedrows = [(yr, cat, dept, round(v[yr])) for (cat, dept), v in fed.items()
               for yr in years if round(v[yr])]
    return brows, frows, fedrows


# --- revenue PDF parsing ----------------------------------------------------
_SOURCES = ["Business Taxes", "Meals & Rentals Tax", "Tobacco Tax",
    "Transfer from Liquor Commission", "Interest & Dividends Tax", "Insurance Tax",
    "Communications Tax", "Real Estate Transfer Tax", "Court Fines & Fees",
    "Securities Revenue", "Beer Tax", "Other", "Video Lottery Terminal",
    "Transfer from Lottery Commission", "Tobacco Settlement", "Utility Property Tax",
    "State Property Tax", "DHHS Recoveries", "Amnesty Payments"]

def _num(t):
    t = t.strip().replace('$', '').replace(',', '').replace('%', '')
    if t in ('', '—', '-'):
        return None
    neg = t.startswith('(') and t.endswith(')')
    t = t.strip('()')
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v

def _nums(s):
    toks = re.findall(r'\(?\$?-?[\d,]+\.\d+\)?|\(?\$?-?[\d,]+\)?%?', s)
    return [x for x in (_num(t) for t in toks) if x is not None]

def _total_group(ns):
    """Rightmost (actual, plan, variance) triple with actual-plan==variance."""
    for i in range(len(ns) - 3, -1, -1):
        a, p, v = ns[i], ns[i + 1], ns[i + 2]
        if abs((a - p) - v) < 0.15:
            return a, p
    if len(ns) >= 2:
        return ns[0], ns[1]
    if len(ns) == 1:
        return ns[0], None
    return None, None

def parse_revenue(pdf_path):
    """{source: (actual_musd, plan_musd)} from page 3 (YTD comparison to plan),
    plus the fiscal year read from the report header."""
    import pdfplumber
    pdf = pdfplumber.open(pdf_path)
    head = pdf.pages[0].extract_text() or ""
    m = re.search(r'FY\s*(\d{2,4})', head)
    yr = int(m.group(1)) if m else None
    if yr and yr < 100:
        yr += 2000
    txt = pdf.pages[2].extract_text() or ""
    for a, b in [("Transfer from Liquor", "Transfer from Liquor Commission"),
                 ("Transfer from Lottery", "Transfer from Lottery Commission")]:
        txt = re.sub(a + r'\s*\n', b + ' ', txt)
    lines = txt.splitlines()
    out = {}
    for name in _SOURCES:
        for l in lines:
            if l.strip().startswith(name):
                a, p = _total_group(_nums(l[len(name):]))
                if a is not None:
                    out[name] = (a, p)
                break
    return yr, out


def main():
    SRC.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    # ---- budget ----
    xlsx = sorted(glob.glob(str(SRC / "*HB*1*.xlsx")) or glob.glob(str(SRC / "*.xlsx")))
    if not xlsx:
        print(f"  (no HB 1 Excel in {SRC}; skipping budget)")
        brows = frows = []
    else:
        brows, frows, fedrows = parse_budget(xlsx[0])
        with open(PROCESSED_DIR / "nh_state_budget.csv", "w", newline="") as f:
            w = csv.writer(f); w.writerow(["fiscal_year", "category", "department", "amount"]); w.writerows(brows)
        with open(PROCESSED_DIR / "nh_state_funding.csv", "w", newline="") as f:
            w = csv.writer(f); w.writerow(["fiscal_year", "source", "amount"]); w.writerows(frows)
        with open(PROCESSED_DIR / "nh_state_federal_funds.csv", "w", newline="") as f:
            w = csv.writer(f); w.writerow(["fiscal_year", "category", "department", "amount"]); w.writerows(fedrows)
        print(f"  budget: {len(brows)} dept-year rows, {len(frows)} funding rows, "
              f"{len(fedrows)} federal-by-agency rows from {Path(xlsx[0]).name}")

    # ---- revenue ----
    pdfs = sorted(glob.glob(str(SRC / "*Monthly_Revenue_June.pdf")) or
                  glob.glob(str(SRC / "*Monthly_Revenue*.pdf")))
    rrows = []
    for p in pdfs:
        yr, d = parse_revenue(p)
        if not yr:
            print(f"  (couldn't read fiscal year from {Path(p).name}; skipping)"); continue
        for s in _SOURCES:
            if s in d:
                a, pl = d[s]
                rrows.append((yr, s, round(a, 1), "" if pl is None else round(pl, 1)))
        print(f"  revenue: FY{yr} from {Path(p).name}")
    if rrows:
        with open(PROCESSED_DIR / "nh_state_revenue.csv", "w", newline="") as f:
            w = csv.writer(f); w.writerow(["fiscal_year", "source", "actual_musd", "plan_musd"]); w.writerows(rrows)
        print(f"  revenue: {len(rrows)} source-year rows written")
    print("state-fiscal: done. Run 'nhbot load' to ingest.")


if __name__ == "__main__":
    main()
