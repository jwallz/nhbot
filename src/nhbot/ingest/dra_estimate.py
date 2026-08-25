#!/usr/bin/env python3
"""
NHbot Phase 0 spike -- compute 2025 equalized (full-value) property tax rates
for all NH municipalities from authoritative NH DRA workbooks.

Run from the repo root (expects the four source workbooks in ./raw/2025/):

    python3 phase0/compute_equalized_rates.py

Sources (NH Dept. of Revenue Administration), tax year 2025:
  A. 2025-municipal-and-village-district-tax-rates.xlsx  -> advertised rate + 4-way split
  B. ratio-median-ratio-cod-prd-ten-year-history.xlsx    -> equalization ratio (weighted mean)
  C. 2025-tables-by-county.xlsx                           -> net assessed valuation + commitment

Equalized rate is computed TWO ways so the methodology gap is explicit:
  simple    = total_advertised_rate * (ratio/100)
  rigorous  = net_tax_commitment / (net_valuation / (ratio/100)) * 1000

Neither is yet validated against DRA's own published "full value tax rate"
(DRA's 2025 comparison is not published as of 2026-08; validate against the
2024 comparison before locking a canonical column). See PHASE0_FINDINGS.md.

Dependencies: openpyxl
"""
import openpyxl, re, csv, os

from nhbot.config import RAW_DIR, PROCESSED_DIR
RAW  = str(RAW_DIR / "2025")
OUT  = str(PROCESSED_DIR)
RETRIEVED = "2026-08-21"
VINTAGE   = 2025
BASE = "https://www.revenue.nh.gov/sites/g/files/ehbemt736/files/documents/"

SRC = {
 "rates":  ("2025-municipal-and-village-district-tax-rates.xlsx",
            BASE + "2025-municipal-and-village-district-tax-rates.xlsx"),
 "ratio":  ("ratio-median-ratio-cod-prd-ten-year-history.xlsx",
            BASE + "ratio-median-ratio-cod-prd-ten-year-history.xlsx"),
 "county": ("2025-tables-by-county.xlsx",
            BASE + "2025-tables-by-county.xlsx"),
}

# 13 NH cities; everything else incorporated is a town.
CITIES = {"Berlin", "Claremont", "Concord", "Dover", "Franklin", "Keene",
          "Laconia", "Lebanon", "Manchester", "Nashua", "Portsmouth",
          "Rochester", "Somersworth"}


def norm(s):
    if s is None:
        return None
    return re.sub(r"\s+", " ", str(s)).strip()


def canon(name):
    """Canonical municipality key: drop the '(U)' tag, fix known name variants
    so the three workbooks join cleanly."""
    if name is None:
        return None
    n = re.sub(r"\s*\(U\)\s*$", "", name).strip()
    variants = {
        "Atkinson & Gilmanton Academy Grant": "Atkinson & Gilmanton",
        "Wentworth Location": "Wentworth's Location",
    }
    return variants.get(n, n)


def is_unincorporated(raw_name):
    return raw_name.strip().endswith("(U)")


def load_rates():
    wb = openpyxl.load_workbook(os.path.join(RAW, SRC["rates"][0]),
                               read_only=True, data_only=True)
    ws = wb["2025 Municipal Tax Rates"]
    out = {}
    for r in ws.iter_rows(min_row=6, values_only=True):
        raw = norm(r[0])
        if not raw:
            continue
        low = raw.lower()
        if low.startswith(("total", "source", "note", "the ", "municipal tax",
                           "new hampshire", "department", "revenue")):
            continue
        if r[8] is None and r[4] is None:
            continue
        # cols: 0 Municipality,1 Date,2 Valuation,3 Val incl utilities,
        # 4 Municipal,5 County,6 State Ed,7 Local Ed,8 Total,9 Commitment
        out[canon(raw)] = dict(
            raw_name=raw, val=r[2], val_util=r[3], mun=r[4], cnty=r[5],
            state_ed=r[6], local_ed=r[7], total=r[8], commitment=r[9],
            unincorporated=is_unincorporated(raw))
    wb.close()
    return out


def load_ratio():
    wb = openpyxl.load_workbook(os.path.join(RAW, SRC["ratio"][0]),
                               read_only=True, data_only=True)
    ws = wb["2025-2016 Summary"]
    out = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        t = canon(norm(r[0]))
        if not t:
            continue
        out[t] = r[2]   # col index 2 == 2025 weighted-mean ratio (percent)
    wb.close()
    return out


def load_county():
    wb = openpyxl.load_workbook(os.path.join(RAW, SRC["county"][0]),
                               read_only=True, data_only=True)
    ws = wb["Alpha Order"]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        t = canon(norm(r[0]))
        if not t:
            continue
        # col 24 Net Valuation, 27 Net Tax Commitment, 28 2025 Tax Rate
        out[t] = dict(net_val=r[24], commitment=r[27], rate=r[28])
    wb.close()
    return out


def classify(key, d):
    if key == "Penacook":            # a village district within Concord, not a town
        return "village_district"
    if d["unincorporated"]:
        return "unincorporated"
    if key in CITIES:
        return "city"
    return "town"


def main():
    rates, ratio, county = load_rates(), load_ratio(), load_county()

    rows = []
    for key, d in rates.items():
        rat = ratio.get(key)
        cty = county.get(key, {})
        total = d["total"]
        eq_simple = round(total * rat / 100, 2) if (total is not None and rat) else None
        net_val, commit = cty.get("net_val"), cty.get("commitment")
        eq_val = (net_val / (rat / 100)) if (net_val and rat) else None
        eq_rig = round(commit / eq_val * 1000, 2) if (eq_val and commit) else None
        rows.append(dict(
            municipality=key, entity_type=classify(key, d),
            municipal_rate=d["mun"], county_rate=d["cnty"],
            local_ed_rate=d["local_ed"], state_ed_rate=d["state_ed"],
            total_rate=total, equalization_ratio_pct=rat,
            # NOT DRA-official: DRA has not published its 2025 full-value
            # comparison yet. These are ESTIMATES (calibrated vs DRA 2024:
            # 'estimate' carries ~+0.13 upward bias, median |err| 0.085).
            # Replace with DRA official when the 2025 comparison publishes.
            equalized_rate_status="ESTIMATE (DRA 2025 official not yet published)",
            equalized_rate_estimate=eq_simple,               # total_rate * ratio
            equalized_rate_estimate_rigorous=eq_rig,         # commitment / equalized_val
            estimate_method="total_rate * equalization_ratio; +0.13 bias, median|err| 0.085 vs DRA 2024",
            net_assessed_valuation=net_val, net_tax_commitment=commit,
            equalized_valuation=round(eq_val) if eq_val else None,
            source_rates=SRC["rates"][1], source_ratio=SRC["ratio"][1],
            source_county=SRC["county"][1],
            retrieved=RETRIEVED, vintage=VINTAGE))

    rows.sort(key=lambda x: x["municipality"])
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, "nh_2025_equalized_rates.csv")
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    munis = [r for r in rows if r["entity_type"] in ("city", "town")]
    uninc = [r for r in rows if r["entity_type"] == "unincorporated"]
    print(f"wrote {path}")
    print(f"rows={len(rows)}  municipalities={len(munis)} "
          f"(cities={sum(1 for r in munis if r['entity_type']=='city')}, "
          f"towns={sum(1 for r in munis if r['entity_type']=='town')})  "
          f"unincorporated={len(uninc)}")


if __name__ == "__main__":
    main()
