#!/usr/bin/env python3
"""
NHbot Phase 0 validation -- calibrate the equalized-rate methodology against
DRA's OWN published 2024 "Comparison of Full Value Tax Rates".

Inputs (in raw/):
  2024/2024-municipal-and-village-tax-rates.xlsx            advertised rate + commitment
  2024/2024-comparison-of-full-value-tax-rates-ranking-order.pdf   DRA official full-value rate
  2025/ratio-median-ratio-cod-prd-ten-year-history.xlsx    equalization ratio (2024 column)

Outputs (phase0/):
  nh_2024_equalized_rates_DRA_official.csv   authoritative -- DRA's published number + equalized valuation
  validation_2024_method_comparison.csv      per-town: DRA official vs simple vs rigorous + errors

DRA's documented definition (from the PDF):
  full_value_rate = gross local property taxes to be raised
                    / total equalized valuation (incl. utility values + equalized railroad)  * 1000
"""
import openpyxl, re, csv, os, statistics as st
import pdfplumber

from nhbot.config import RAW_DIR, PROCESSED_DIR
RAW  = str(RAW_DIR)
OUT  = str(PROCESSED_DIR)
CITIES = {"Berlin","Claremont","Concord","Dover","Franklin","Keene","Laconia",
          "Lebanon","Manchester","Nashua","Portsmouth","Rochester","Somersworth"}

def norm(s): return re.sub(r"\s+"," ",str(s)).strip() if s is not None else None
def canon(n):
    if n is None: return None
    n = re.sub(r"\s*\(U\)\s*$","",n).strip()
    return {"Atkinson & Gilmanton Academy Grant":"Atkinson & Gilmanton",
            "Wentworth Location":"Wentworth's Location"}.get(n,n)
def fnum(s):
    if s is None: return None
    s = str(s).replace(",","").strip()
    if s in ("","N/A"): return None
    try: return float(s)
    except: return None

def load_rates():
    wb = openpyxl.load_workbook(f"{RAW}/2024/2024-municipal-and-village-tax-rates.xlsx",
                               read_only=True, data_only=True)
    ws = wb["2024 Municipal Tax Rates"]; out={}
    for r in ws.iter_rows(min_row=6, values_only=True):
        m = norm(r[0])
        if not m: continue
        if m.lower().startswith(("total","source","note","the ","municipal tax",
                                 "new hampshire","department","revenue")): continue
        if r[8] is None and r[4] is None: continue
        out[canon(m)] = dict(val=r[2], val_util=r[3], total=r[8], commit=r[9])
    wb.close(); return out

def load_ratio_2024():
    wb = openpyxl.load_workbook(f"{RAW}/2025/ratio-median-ratio-cod-prd-ten-year-history.xlsx",
                               read_only=True, data_only=True)
    ws = wb["2025-2016 Summary"]; out={}
    for r in ws.iter_rows(min_row=3, values_only=True):
        t = canon(norm(r[0]))
        if t: out[t] = r[7]   # col idx 7 == 2024 ratio
    wb.close(); return out

def load_pub():
    pdf = pdfplumber.open(f"{RAW}/2024/2024-comparison-of-full-value-tax-rates-ranking-order.pdf")
    out={}
    for pi in range(2, len(pdf.pages)):
        for t in pdf.pages[pi].extract_tables():
            for row in t:
                if not row or not row[0]: continue
                name = canon(re.sub(r"\s+"," ",str(row[0])).strip())
                if name.lower().startswith(("municipality","2024","average","nh department","municipal and")): continue
                if len(row) < 7: continue
                lr, fv = fnum(row[3]), fnum(row[5])
                if lr is None and fv is None: continue
                out[name] = dict(mod_assessed=fnum(row[1]), equalized_val=fnum(row[2]),
                                 local_rate=lr, dra_ratio=fnum(row[4]),
                                 full_value_rate=fv, rank=str(row[6]).strip())
    pdf.close(); return out

def etype(k):
    return "city" if k in CITIES else "town"

def main():
    rates, ratio, pub = load_rates(), load_ratio_2024(), load_pub()
    os.makedirs(OUT, exist_ok=True)

    # 1) authoritative DRA official dataset
    official=[]
    for k,p in sorted(pub.items()):
        if p["full_value_rate"] in (None,0) and p["local_rate"] in (None,0): continue
        official.append(dict(municipality=k, full_value_rate_official=p["full_value_rate"],
                             local_total_rate=p["local_rate"], equalization_ratio=p["dra_ratio"],
                             equalized_valuation_incl_util_rr=p["equalized_val"],
                             modified_local_assessed_value=p["mod_assessed"], rank=p["rank"],
                             vintage=2024, source="2024-comparison-of-full-value-tax-rates-ranking-order.pdf"))
    with open(f"{OUT}/nh_2024_equalized_rates_DRA_official.csv","w",newline="") as f:
        w=csv.DictWriter(f, fieldnames=list(official[0].keys())); w.writeheader(); w.writerows(official)

    # 2) method comparison for incorporated municipalities
    comp=[]
    for k,d in rates.items():
        if k=="Penacook": continue
        p=pub.get(k); rat=ratio.get(k)
        if not p or p["full_value_rate"] in (None,0) or not rat: continue
        dra=p["full_value_rate"]
        simple=round(d["total"]*rat/100,2)
        rigorous=round(d["commit"]/(d["val_util"]/(rat/100))*1000,2)
        comp.append(dict(municipality=k, entity_type=etype(k), dra_official=dra,
                         simple=simple, simple_err=round(simple-dra,2),
                         rigorous=rigorous, rigorous_err=round(rigorous-dra,2),
                         ratio_history=rat, ratio_dra=p["dra_ratio"],
                         ratio_source_mismatch=(abs(rat-p["dra_ratio"])>1 if p["dra_ratio"] else None)))
    comp.sort(key=lambda r:-abs(r["simple_err"]))
    with open(f"{OUT}/validation_2024_method_comparison.csv","w",newline="") as f:
        w=csv.DictWriter(f, fieldnames=list(comp[0].keys())); w.writeheader(); w.writerows(comp)

    def stats(key):
        e=[r[key] for r in comp]; ae=[abs(x) for x in e]
        return (st.mean(e), st.mean(ae), st.median(ae), max(ae),
                sum(1 for x in ae if x<=0.10), sum(1 for x in ae if x<=0.25))
    print(f"compared {len(comp)} municipalities; DRA official rows {len(official)}")
    for m in ("simple_err","rigorous_err"):
        mn,ma,md,mx,w10,w25=stats(m)
        print(f"{m:12s} bias={mn:+.3f} mean|e|={ma:.3f} median|e|={md:.3f} max|e|={mx:.2f} "
              f"within0.10={w10}/{len(comp)} within0.25={w25}")
    mism=[r for r in comp if r["ratio_source_mismatch"]]
    print("ratio-source mismatches (>1 pt):", [(r["municipality"],r["ratio_history"],r["ratio_dra"]) for r in mism])

if __name__=="__main__":
    main()
